import itertools, openpyxl
from functools import reduce
from openpyxl.styles import Alignment, PatternFill, Side, Border

def printError(alphabet, xor, parities, newData):
    setColors = {}

    i = 0
    for k in xor.keys():
        setColors[k] = colors[i]
        i += 1

    index = 0
    for k, v in xor.items():
        text = 'Pos ' + k[1:] + ': '
        for i in parities[k]:
            text += str(newData[bottom[i]]) + ' XOR '
        
        text = text[:len(text) - 4] + '= ' + str(v)
        sheet[alphabet[10 + p] + str(16 + index)] = text
        sheet[alphabet[10 + p] + str(16 + index)].fill = setColors[k]
        index += 1

    index = 0
    for k in reversed(xor.keys()):
        sheet[alphabet[12 + p + index] + str(16)] = 'Pos ' + k[1:]
        currentCell = sheet.cell(row = 16, column = 13 + p + index)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.fill = setColors[k]
        currentCell.border = border

        sheet[alphabet[12 + p + index] + str(17)] = xor[k]
        currentCell = sheet.cell(row = 17, column = 13 + p + index)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.fill = setColors[k]
        currentCell.border = border

        index += 1

def printParities(alphabet, parities, values, newData):
    index = 0
    for k, v in parities.items():
        text = k + ' = '
        for i in v:
            text += str(i) + ', '
        text = text[:len(text) - 2]

        sheet[alphabet[4 + p] + str(16 + index)] = text
        index += 1

    index = 0
    for k, v in parities.items():
        text = k + ' + '
        for i in range(1, len(v)):
            text += str(newData[bottom[v[i]]]) + ' + '

        text = text[:len(text) - 3]

        sheet[alphabet[6 + p] + str(16 + index)] = text
        index += 1


    index = 0
    for k, v in values.items():
        text = '-> ' + k + ' = ' + str(v)
        sheet[alphabet[8 + p] + str(16 + index)] = text
        sheet[alphabet[8 + p] + str(16 + index)].border = border
        sheet[alphabet[8 + p] + str(16 + index)].fill = green
        index += 1

def printTruthTable(alphabet, truthTable):
    for i in range(p):
        sheet[alphabet[3 + i] + '14'] = f'p{p - 1 - i}'
        currentCell = sheet.cell(row = 14, column = 4 + i)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.fill = green
        currentCell.border = border

    for i in range(len(truthTable)):
        sheet[alphabet[2] + str(15 + i)] = i
        currentCell = sheet.cell(row = 15 + i, column = 3)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.fill = orange
        currentCell.border = border
        if i > p + d:
            currentCell.fill = black
        for j in range(p):
            sheet[alphabet[3 + j] + str(15 + i)] = truthTable[i][j]
            currentCell = sheet.cell(row = 15 + i, column = 4 + j)
            currentCell.alignment = Alignment(horizontal='center')
            if truthTable[i][j] == 1:
                currentCell.fill = lightBlue

            if i > p + d:
                currentCell.fill = black

            currentCell.border = border


def printAllData(alphabet, newData, top, bottom, error):
    parityIndexes = []

    for k, v in top.items():
        if k[0] == 'p':
            parityIndexes.append(v)

        sheet[alphabet[3 + v] + '9'] = k
        currentCell = sheet.cell(row = 9, column = 4 + v)
        currentCell.alignment = Alignment(horizontal='center')

    for i in range(len(newData)):
        sheet[alphabet[3 + i] + '10'] = newData[i]
        currentCell = sheet.cell(row = 10, column = 4 + i)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.fill = yellow
        if i in parityIndexes:
            currentCell.fill = green

        if error[0] == True and error[1] == i:
            currentCell.fill = red

        currentCell.border = border

    for k, v in bottom.items():
        sheet[alphabet[3 + v] + '11'] = k
        currentCell = sheet.cell(row = 11, column = 4 + v)
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.fill = gray
        currentCell.border = border

def calculateParityBits(d):
    for i in range(d):
        if 2 ** i >= d + i + 1:
            return i

def posParityBits(data, p):
    d = len(data)
    newData = []
    top = {}
    bottom = {i:abs(i - d - p) for i in range(9, 0, -1)}

    dataIndex = d - 1
    parityIndex = 0

    for i in range(1, d + p + 1):
        if i == 2 ** parityIndex:
            newData.append(0)
            top['p' + str(parityIndex)] = abs(i - d - p)
            parityIndex += 1
        else:
            newData.append(int(data[dataIndex]))
            top['d' + str(abs(dataIndex - d + 1))] = abs(i - d - p)
            dataIndex -= 1

    top = {k: top[k] for k in reversed((top.keys()))}
    return top, newData[::-1], bottom

def getParityBits(newData, p, truthTable, bottom):
    newD = len(newData)
    parities = {}

    for i in range(p - 1, -1, -1):
        indexes = []
        for j in range(newD + 1):
            if truthTable[j][i] == 1:
                indexes.append(j)
        parities['p' + str(abs(i - p + 1))] = indexes
    
    values = {}

    for k, v in parities.items():
        total = 0
        for i in v:
            total += newData[bottom[i]]

        if total % 2 == 0:
            values[k] = 0
        else:
            values[k] = 1
    
    return parities, values

def updateNewData(newData, top, values):
    for k, v in values.items():
        newData[top[k]] = v

def detectError(newData, bottom, parities):
    errors = {}
    for k, v in parities.items():
        values = [int(newData[bottom[i]]) for i in v]
        xor = reduce(lambda x, y: x ^ y, values)
        errors[k] = xor
    
    binary = ''
    for _, v in errors.items():
        binary += str(v)
    
    binary = binary[::-1]
    decimal = int(binary, 2)

    return errors, decimal

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

purple = PatternFill("solid", fgColor="FF00F6")
yellow = PatternFill("solid", fgColor="FFFF00")
green = PatternFill("solid", fgColor="50FF00")
orange = PatternFill("solid", fgColor="FF8900")
lightBlue = PatternFill("solid", fgColor="00F8FF")
red = PatternFill("solid", fgColor="FF0000")
gray = PatternFill("solid", fgColor="A2C4C9")
black = PatternFill("solid", fgColor="000000")

colors = [
    PatternFill("solid", fgColor="FFD966"),
    PatternFill("solid", fgColor="E69138"),
    PatternFill("solid", fgColor="93C47D"),
    PatternFill("solid", fgColor="A2C4C9"),
    PatternFill("solid", fgColor="C6E39F")
]

thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

fileName = 'hamming.xlsx'
sheetName = 'Hoja1'

workbook = openpyxl.load_workbook(fileName)
#sheet = workbook.active

idx = workbook.sheetnames.index(sheetName)
sheet = workbook[sheetName]
workbook.remove(sheet)
workbook.create_sheet(sheetName, idx)
sheet = workbook[sheetName]

for row in sheet['A1:Z100']:
  for cell in row:
    cell.value = None

sheet['B2'] = 'p = Número de bits de paridad'
sheet['B3'] = 'd = Número de bits de datos'
sheet['B6'] = 'bits de datos'

#data = '10110'
data = input('Ingrese real: ')

for i in range(len(data)):
    cell = alphabet[3 + i] + '6'
    sheet[cell] = int(data[i])
    currentCell = sheet.cell(row = 6, column = 4 + i)
    currentCell.alignment = Alignment(horizontal='center')
    currentCell.fill = yellow
    currentCell.border = border


d = len(data)
p = calculateParityBits(d)

sheet['B9'] = f'd = {d}'
sheet['B10'] = f'p = {p}'
sheet['B9'].fill = purple
sheet['B9'].border = border
sheet['B10'].fill = purple
sheet['B10'].border = border

top, newData, bottom = posParityBits(data, p)

truthTable = list(itertools.product([int(False), int(True)], repeat = p))

parities, values = getParityBits(newData, p, truthTable, bottom)

updateNewData(newData, top, values)

#errorData = '10100'
errorData = input('Ingrese erroneo: ')
top, errorNewData, bottom = posParityBits(errorData, p)
updateNewData(errorNewData, top, values)
xor, errorIndex = detectError(errorNewData, bottom, parities)

printTruthTable(alphabet, truthTable)

printParities(alphabet, parities, values, newData)

printError(alphabet, xor, parities, errorNewData)

error = ()

if errorIndex == 0:
    print('No hay error')
    sheet[alphabet[10 + p + p - 2] + str(19)] = 'No hay error'
    error = (False, 0)
else:
    print('El error se encuentra en la posicion', errorIndex)
    sheet[alphabet[10 + p + p - 2] + str(19)] = 'Error en el índice: ' + str(errorIndex)
    error = (True, bottom[errorIndex])

printAllData(alphabet, errorNewData, top, bottom, error)

workbook.save(fileName)